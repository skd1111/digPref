"""Phase 7 V0 · Excel 导出 —— openpyxl 多 Sheet + 条件格式 + 列宽。

design §4.3：
  - 多 Sheet、表头合并、条件格式（负数标红）、自动列宽
  - 中文无乱码（openpyxl 原生 UTF-8）
"""
from __future__ import annotations

import hashlib
import tempfile
from pathlib import Path
from typing import Any

from agent.dataexpert.export.watermark import embed_watermark_metadata, mask_pii_columns


def export_excel(
    columns: list[str],
    rows: list[list[Any]],
    *,
    title: str = "数据报表",
    operator: str = "current_user",
    output_path: str | None = None,
) -> dict[str, Any]:
    """导出 Excel 文件。

    Args:
        columns: 列名。
        rows: 数据行。
        title: 报表标题。
        operator: 操作人（水印用）。
        output_path: 输出路径（None 则用临时文件）。

    Returns:
        {"path": str, "md5": str, "row_count": int, "watermark": str}
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, numbers
        from openpyxl.utils import get_column_letter
    except ImportError:
        # openpyxl 未安装时降级为 CSV
        return _fallback_csv(columns, rows, title, operator, output_path)

    # PII 脱敏
    masked_rows = mask_pii_columns(columns, rows)

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet 名最长 31 字符

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

    # 写表头
    for ci, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 写数据 + 条件格式（负数标红）
    red_font = Font(color="FF0000")
    for ri, row in enumerate(masked_rows, 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            # 负数标红
            if isinstance(val, (int, float)) and val < 0:
                cell.font = red_font

    # 自动列宽
    for ci in range(1, len(columns) + 1):
        max_len = len(str(columns[ci - 1]))
        for ri in range(2, min(len(masked_rows) + 2, 102)):  # 采样前 100 行
            cell_val = ws.cell(row=ri, column=ci).value
            if cell_val is not None:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 50)

    # 保存
    if output_path is None:
        output_path = str(Path(tempfile.gettempdir()) / f"eaide_export_{title}.xlsx")
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    # 计算 MD5
    file_bytes = Path(output_path).read_bytes()
    md5 = hashlib.md5(file_bytes).hexdigest()

    # 水印元数据
    meta = embed_watermark_metadata({
        "path": output_path,
        "md5": md5,
        "row_count": len(masked_rows),
        "format": "excel",
    }, operator)

    return meta


def _fallback_csv(
    columns: list[str], rows: list[list[Any]],
    title: str, operator: str, output_path: str | None,
) -> dict[str, Any]:
    """openpyxl 不可用时降级为 CSV。"""
    from agent.dataexpert.export.csv import export_csv
    return export_csv(columns, rows, title=title, operator=operator, output_path=output_path)
